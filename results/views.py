from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status, permissions
from django.db.models import Q, Avg
from django.core.paginator import Paginator
from django.shortcuts import get_object_or_404
from django.db import connection
from .models import Result
from .serializers import ResultSerializer, ResultDetailSerializer
from tests.models import Test
from datetime import datetime, timedelta
from core.permissions import IsAdmin, IsOwnerOrAdmin
import logging

logger = logging.getLogger(__name__)

@api_view(['GET', 'POST'])
@permission_classes([permissions.IsAuthenticated])
def result_list(request):
    """
    GET: List all results (with filters and pagination)
    POST: Create new result (Usually auto-created when test is submitted)
    """
    if request.method == 'GET':
        if request.user.role == 'admin':
            queryset = Result.objects.all()
        else:
            queryset = Result.objects.filter(candidate=request.user)

        passed = request.query_params.get('passed')
        if passed:
            queryset = queryset.filter(passed=(passed.lower() == 'true'))
        
        candidate_id = request.query_params.get('candidate')
        if candidate_id and request.user.role == 'admin':
            queryset = queryset.filter(candidate_id=candidate_id)

        search = request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(candidate__name__icontains=search) |
                Q(candidate__email__icontains=search) |
                Q(test__test_id__icontains=search)
            )
        
        from_date = request.query_params.get('from_date')
        if from_date:
            queryset = queryset.filter(evaluated_at__gte=from_date)
        
        to_date = request.query_params.get('to_date')
        if to_date:
            queryset = queryset.filter(evaluated_at__lte=to_date)

        ordering = request.query_params.get('ordering', '-evaluated_at')
        queryset = queryset.order_by(ordering)
        
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 10))
        
        paginator = Paginator(queryset, page_size)
        current_page = paginator.get_page(page)
        
        serializer = ResultSerializer(current_page, many=True)
        
        return Response({
            'success': True,
            'data': {
                'count': paginator.count,
                'total_pages': paginator.num_pages,
                'current_page': page,
                'page_size': page_size,
                'has_next': current_page.has_next(),
                'has_previous': current_page.has_previous(),
                'results': serializer.data
            }
        })
    
    elif request.method == 'POST':
        if request.user.role != 'admin':
            return Response({
                'success': False,
                'message': 'Admin permission required'
            }, status=status.HTTP_403_FORBIDDEN)
        
        from pymongo import MongoClient
        from django.conf import settings
        from bson import ObjectId
        from bson.errors import InvalidId
        from datetime import datetime
        
        data = request.data
        test_id = data.get('test')
        candidate_id = data.get('candidate')
        
        try:
            client = MongoClient(settings.DATABASES['default']['CLIENT']['host'])
            db = client[settings.DATABASES['default']['NAME']]
        except Exception as e:
            return Response({
                'success': False,
                'message': f'Database connection error: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        test_exists = False
        test_collection = db['tests']

        test_doc = test_collection.find_one({'_id': test_id})
        if test_doc:
            test_exists = True

        if not test_exists:
            try:
                test_doc = test_collection.find_one({'_id': ObjectId(test_id)})
                if test_doc:
                    test_exists = True
            except InvalidId:
                pass
        
        if not test_exists:
            return Response({
                'success': False,
                'errors': {
                    'test': [f'Test with ID {test_id} does not exist in database.']
                }
            }, status=status.HTTP_400_BAD_REQUEST)
        
        candidate_exists = False
        users_collection = db['users']
        
        candidate_doc = users_collection.find_one({'_id': candidate_id})
        if candidate_doc:
            candidate_exists = True
        
        if not candidate_exists:
            try:
                candidate_doc = users_collection.find_one({'_id': ObjectId(candidate_id)})
                if candidate_doc:
                    candidate_exists = True
            except InvalidId:
                pass
        
        if not candidate_exists:
            return Response({
                'success': False,
                'errors': {
                    'candidate': [f'Candidate with ID {candidate_id} does not exist in database.']
                }
            }, status=status.HTTP_400_BAD_REQUEST)

        existing_result = None
        
        existing_result = db['results'].find_one({
            'test_id': test_id,
            'candidate_id': candidate_id
        })
        
        if not existing_result:
            existing_result = db['results'].find_one({
                'test': test_id,
                'candidate': candidate_id
            })
        
        if existing_result:
            return Response({
                'success': False,
                'message': 'Result already exists for this test and candidate'
            }, status=status.HTTP_400_BAD_REQUEST)

        result_doc = {
            'test_id': test_id,
            'candidate_id': candidate_id,

            'test': test_id,
            'candidate': candidate_id,

            'total_marks': data.get('total_marks', 0),
            'obtained_marks': data.get('obtained_marks', 0),
            'percentage': data.get('percentage', 0),
            'passed': data.get('passed', False),
            'answers': data.get('answers', {}),
            'evaluated_at': datetime.now(),
            'created_at': datetime.now(),
            'updated_at': datetime.now()
        }

        optional_fields = ['total_questions', 'attempted', 'correct', 'wrong', 'skipped']
        for field in optional_fields:
            if field in data:
                result_doc[field] = data[field]

        try:
            result_id = db['results'].insert_one(result_doc).inserted_id

            created_result = db['results'].find_one({'_id': result_id})

            created_result['_id'] = str(created_result['_id'])
            
            return Response({
                'success': True,
                'message': 'Result created successfully',
                'data': created_result
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            error_msg = str(e)

            if 'duplicate key error' in error_msg:
                try:
                    db['results'].drop_index('test_1')
                    result_id = db['results'].insert_one(result_doc).inserted_id
                    
                    return Response({
                        'success': True,
                        'message': 'Result created successfully after dropping index',
                        'data': {'id': str(result_id)}
                    }, status=status.HTTP_201_CREATED)
                except:
                    return Response({
                        'success': False,
                        'message': 'Duplicate key error. Please use a different test/candidate combination.'
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            return Response({
                'success': False,
                'message': f'Error creating result: {error_msg}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET', 'PUT', 'PATCH', 'DELETE'])
@permission_classes([permissions.IsAuthenticated, IsOwnerOrAdmin])
def result_detail(request, pk):
    """
    GET: Retrieve result details
    PUT/PATCH: Update result (Admin only)
    DELETE: Delete result (Admin only)
    """
    try:
        result = Result.objects.get(pk=pk)
        
        if request.user.role != 'admin' and result.candidate != request.user:
            return Response({
                'success': False,
                'message': 'You do not have permission to access this result'
            }, status=status.HTTP_403_FORBIDDEN)
            
    except Result.DoesNotExist:
        return Response({
            'success': False,
            'message': 'Result not found'
        }, status=status.HTTP_404_NOT_FOUND)
    
    if request.method == 'GET':
        serializer = ResultDetailSerializer(result)
        return Response({
            'success': True,
            'data': serializer.data
        })
    
    if request.user.role != 'admin':
        return Response({
            'success': False,
            'message': 'Admin permission required'
        }, status=status.HTTP_403_FORBIDDEN)
    
    if request.method in ['PUT', 'PATCH']:
        serializer = ResultSerializer(result, data=request.data, partial=(request.method == 'PATCH'))
        if serializer.is_valid():
            serializer.save()
            return Response({
                'success': True,
                'message': 'Result updated successfully',
                'data': serializer.data
            })
        return Response({
            'success': False,
            'errors': serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'DELETE':
        result.delete()
        return Response({
            'success': True,
            'message': 'Result deleted successfully'
        }, status=status.HTTP_204_NO_CONTENT)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def result_my_results(request):
    """Get current user's results"""
    results = Result.objects.filter(candidate=request.user).order_by('-evaluated_at')
    
    page = int(request.query_params.get('page', 1))
    page_size = int(request.query_params.get('page_size', 10))
    
    paginator = Paginator(results, page_size)
    current_page = paginator.get_page(page)
    
    serializer = ResultSerializer(current_page, many=True)
    
    return Response({
        'success': True,
        'data': {
            'count': paginator.count,
            'total_pages': paginator.num_pages,
            'current_page': page,
            'page_size': page_size,
            'has_next': current_page.has_next(),
            'has_previous': current_page.has_previous(),
            'results': serializer.data
        }
    })


@api_view(['GET'])
@permission_classes([IsAdmin])
def result_summary(request):
    """Get results summary (Admin only)"""
    
    try:
        from pymongo import MongoClient
        from django.conf import settings
        from datetime import datetime, timedelta
        import logging
        
        logger = logging.getLogger(__name__)
        
        client = MongoClient(settings.DATABASES['default']['CLIENT']['host'])
        db = client[settings.DATABASES['default']['NAME']]
        collection = db['results']

        total_tests = collection.count_documents({})
        passed_tests = collection.count_documents({'passed': True})
        failed_tests = collection.count_documents({'passed': False})

        pipeline = [
            {"$group": {
                "_id": None,
                "avg_percentage": {"$avg": "$percentage"}
            }}
        ]
        result = list(collection.aggregate(pipeline))
        avg_score = result[0]['avg_percentage'] if result else 0
        
        level_stats = {}
        
        all_results = list(collection.find({}))
        
        for res in all_results:
            test_id = res.get('test_id') or res.get('test')
            
            if test_id:
                test_collection = db['tests']
                test_doc = test_collection.find_one({'_id': test_id})
                
                if test_doc:
                    level = test_doc.get('experience_level', 'unknown')
                    
                    if level not in level_stats:
                        level_stats[level] = {'total': 0, 'passed': 0, 'total_score': 0}
                    
                    level_stats[level]['total'] += 1
                    if res.get('passed', False):
                        level_stats[level]['passed'] += 1
                    level_stats[level]['total_score'] += res.get('percentage', 0)
        
        for level in level_stats:
            if level_stats[level]['total'] > 0:
                level_stats[level]['avg_score'] = round(
                    level_stats[level]['total_score'] / level_stats[level]['total'], 2
                )
            if 'total_score' in level_stats[level]:
                del level_stats[level]['total_score']
        
        tech_stats = {}
        
        for doc in all_results:
            if doc.get('technology_wise'):
                for tech, stats in doc['technology_wise'].items():
                    if tech not in tech_stats:
                        tech_stats[tech] = {
                            'total_questions': 0, 
                            'correct': 0, 
                            'tests': 0,
                            'total_score': 0
                        }
                    
                    tech_stats[tech]['total_questions'] += stats.get('total', 0)
                    tech_stats[tech]['correct'] += stats.get('correct', 0)
                    tech_stats[tech]['tests'] += 1
                    
                    if stats.get('total', 0) > 0:
                        accuracy = (stats.get('correct', 0) / stats.get('total', 1)) * 100
                        tech_stats[tech]['total_score'] += accuracy
        
        for tech in tech_stats:
            if tech_stats[tech]['tests'] > 0:
                tech_stats[tech]['avg_accuracy'] = round(
                    tech_stats[tech]['total_score'] / tech_stats[tech]['tests'], 2
                )
            if tech_stats[tech]['total_questions'] > 0:
                tech_stats[tech]['overall_accuracy'] = round(
                    (tech_stats[tech]['correct'] / tech_stats[tech]['total_questions']) * 100, 2
                )
            if 'total_score' in tech_stats[tech]:
                del tech_stats[tech]['total_score']
        
        thirty_days_ago = datetime.now() - timedelta(days=30)
        
        recent_results = collection.find({
            'evaluated_at': {'$gte': thirty_days_ago}
        })
        
        daily_stats = {}
        for doc in recent_results:
            evaluated_at = doc.get('evaluated_at')
            if evaluated_at:
                if isinstance(evaluated_at, str):
                    try:
                        date_obj = datetime.fromisoformat(evaluated_at.replace('Z', '+00:00'))
                        date_str = date_obj.strftime('%Y-%m-%d')
                    except:
                        date_str = evaluated_at[:10]
                else:
                    date_str = evaluated_at.strftime('%Y-%m-%d')
                
                if date_str not in daily_stats:
                    daily_stats[date_str] = {'total': 0, 'passed': 0, 'total_score': 0}
                
                daily_stats[date_str]['total'] += 1
                if doc.get('passed', False):
                    daily_stats[date_str]['passed'] += 1
                daily_stats[date_str]['total_score'] += doc.get('percentage', 0)
        
        for date_str in daily_stats:
            if daily_stats[date_str]['total'] > 0:
                daily_stats[date_str]['avg_score'] = round(
                    daily_stats[date_str]['total_score'] / daily_stats[date_str]['total'], 2
                )
                daily_stats[date_str]['pass_rate'] = round(
                    (daily_stats[date_str]['passed'] / daily_stats[date_str]['total']) * 100, 2
                )
            if 'total_score' in daily_stats[date_str]:
                del daily_stats[date_str]['total_score']
        
        sorted_daily = dict(sorted(daily_stats.items()))
        
        return Response({
            'success': True,
            'data': {
                'overall': {
                    'total_tests': total_tests,
                    'passed': passed_tests,
                    'failed': failed_tests,
                    'pass_rate': round((passed_tests / total_tests * 100) if total_tests > 0 else 0, 2),
                    'average_score': round(avg_score, 2)
                },
                'level_wise': level_stats,
                'technology_wise': tech_stats,
                'daily_trends': sorted_daily
            }
        })
        
    except Exception as e:
        logger.error(f"Error in result_summary: {str(e)}")
        import traceback
        traceback.print_exc()
        
        return Response({
            'success': False,
            'message': f'Error generating summary: {str(e)}',
            'data': {
                'overall': {
                    'total_tests': 0,
                    'passed': 0,
                    'failed': 0,
                    'pass_rate': 0,
                    'average_score': 0
                },
                'level_wise': {},
                'technology_wise': {},
                'daily_trends': {}
            }
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAdmin])
def result_export(request):
    """Export results as CSV/JSON"""
    format_type = request.query_params.get('format', 'json')
    
    queryset = Result.objects.all().select_related('test', 'candidate')
    
    from_date = request.query_params.get('from_date')
    if from_date:
        queryset = queryset.filter(evaluated_at__gte=from_date)
    
    to_date = request.query_params.get('to_date')
    if to_date:
        queryset = queryset.filter(evaluated_at__lte=to_date)
    
    passed = request.query_params.get('passed')
    if passed:
        queryset = queryset.filter(passed=(passed.lower() == 'true'))
    
    level = request.query_params.get('level')
    if level:
        queryset = queryset.filter(test__experience_level=level)
    
    data = []
    for result in queryset:
        tech_list = ', '.join(result.test.selected_technologies) if hasattr(result.test, 'selected_technologies') else ''
        
        row = {
            'test_id': result.test.test_id if hasattr(result.test, 'test_id') else '',
            'candidate_name': result.candidate.name if result.candidate else '',
            'candidate_email': result.candidate.email if result.candidate else '',
            'experience_level': result.test.experience_level if hasattr(result.test, 'experience_level') else '',
            'technologies': tech_list,
            'total_questions': result.total_questions,
            'attempted': result.attempted,
            'correct': result.correct,
            'wrong': result.wrong,
            'skipped': result.skipped,
            'percentage': result.percentage,
            'passed': 'Yes' if result.passed else 'No',
            'evaluated_at': result.evaluated_at.isoformat() if result.evaluated_at else ''
        }
        data.append(row)
    
    if format_type == 'csv':
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="results_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        if data:
            writer = csv.DictWriter(response, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        else:
            writer = csv.writer(response)
            writer.writerow(['No data available'])
        
        return response
    
    elif format_type == 'excel':
        try:
            import openpyxl
            from django.http import HttpResponse
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Results Export"
            
            if data:
                headers = list(data[0].keys())
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                
                for row_idx, row_data in enumerate(data, 2):
                    for col_idx, header in enumerate(headers, 1):
                        ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ''))

                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column].width = adjusted_width
            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="results_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            wb.save(response)
            return response
            
        except ImportError:
            return Response({
                'success': False,
                'message': 'openpyxl not installed. Please install it for Excel export.'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    return Response({
        'success': True,
        'data': data
    })


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def result_by_candidate(request, candidate_id):
    """Get results for a specific candidate - Using Raw MongoDB"""
    try:
        from pymongo import MongoClient
        from django.conf import settings
        from bson import ObjectId
        from bson.errors import InvalidId
        from accounts.models import User
        
        candidate = None
        try:
            candidate = User.objects.get(_id=ObjectId(candidate_id))
        except (InvalidId, User.DoesNotExist):
            try:
                candidate = User.objects.get(_id=candidate_id)
            except User.DoesNotExist:
                pass
        
        if not candidate:
            return Response({
                'success': False,
                'message': f'Candidate with ID {candidate_id} does not exist.'
            }, status=status.HTTP_404_NOT_FOUND)
        
        if request.user.role != 'admin' and str(request.user._id) != str(candidate_id):
            return Response({
                'success': False,
                'message': 'Permission denied'
            }, status=status.HTTP_403_FORBIDDEN)
        
        client = MongoClient(settings.DATABASES['default']['CLIENT']['host'])
        db = client[settings.DATABASES['default']['NAME']]
        collection = db['results']
        
        cursor = collection.find({
            '$or': [
                {'candidate_id': candidate_id},
                {'candidate': candidate_id}
            ]
        }).sort('evaluated_at', -1)
        
        all_results = list(cursor)
        
        for result in all_results:
            result['_id'] = str(result['_id'])
        
        total_tests = len(all_results)
        passed_tests = len([r for r in all_results if r.get('passed') == True])
        
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 10))
        
        start = (page - 1) * page_size
        end = start + page_size
        paginated_results = all_results[start:end]
        
        tests_collection = db['tests']
        for result in paginated_results:
            test_id = result.get('test_id') or result.get('test')
            if test_id:
                test = tests_collection.find_one({'_id': test_id})
                if test:
                    result['test_details'] = {
                        'test_id': test.get('test_id'),
                        'experience_level': test.get('experience_level'),
                        'technologies': test.get('selected_technologies')
                    }
        
        return Response({
            'success': True,
            'data': {
                'candidate_id': candidate_id,
                'candidate_name': candidate.name,
                'candidate_email': candidate.email,
                'total_tests': total_tests,
                'passed_tests': passed_tests,
                'failed_tests': total_tests - passed_tests,
                'pass_rate': round((passed_tests / total_tests * 100) if total_tests > 0 else 0, 2),
                'pagination': {
                    'current_page': page,
                    'page_size': page_size,
                    'total_pages': (total_tests + page_size - 1) // page_size if total_tests > 0 else 0,
                    'has_next': end < total_tests,
                    'has_previous': page > 1
                },
                'results': paginated_results
            }
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return Response({
            'success': False,
            'message': f'Error: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def result_by_test(request, test_id):
    """Get result for a specific test"""
    try:
        result = Result.objects.get(test_id=test_id)

        if request.user.role != 'admin' and result.candidate != request.user:
            return Response({
                'success': False,
                'message': 'Permission denied'
            }, status=status.HTTP_403_FORBIDDEN)
        
        serializer = ResultDetailSerializer(result)
        return Response({
            'success': True,
            'data': serializer.data
        })
        
    except Result.DoesNotExist:
        return Response({
            'success': False,
            'message': 'Result not found for this test'
        }, status=status.HTTP_404_NOT_FOUND)